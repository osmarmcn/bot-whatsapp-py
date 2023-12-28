from openpyxl import*
from urllib.parse import quote
import webbrowser
import pyautogui
from time import sleep

webbrowser.open('https://web.whatsapp.com')
sleep(30)

workbook = load_workbook('clientes.xlsx')
pagina_clientes = workbook['Sheet1']


for linha in pagina_clientes.iter_rows(min_row=2): 
    nome = linha[0].value
    telefone = linha[1].value
    pagamento = linha[2].value
    mensagem = f'Olá, {nome} seu boleto está perto do vencimento, {pagamento.strftime('%d/%m/%Y')}'
    link_mensagem = f'https://web.whatsapp.com/send?phone={telefone}&text={quote(mensagem)}'

    
    try:
        webbrowser.open(link_mensagem)
        sleep(10)
        seta = pyautogui.locateCenterOnScreen('seta.png')
        sleep(5)
        pyautogui.click(seta[0], seta[1])
        sleep(5)
        pyautogui.hotkey('ctrl', 'w')
        sleep(5)
    except:
        print(f'Não foi possivel enviar mensagem para {nome}')
        with open('erros.csv', 'a',newline='', encoding='utf-8') as arquivo:
            arquivo.write(f'{nome}, {telefone}')


    # https://web.whatsapp.com/send?phone={numero especifico}&text={mensagem especifica}
    #print(nome, telefone, pagamento)


    















''''
iter_rows() - é uma função do openpyxl que vai passar por todas as linhas da lista.
obs: o parâmetro min_row=2 indica que vai começar na linha dois, tem que vericar na lista onde realmente começa a lista pdoendo ser uma linha diferente.




'''